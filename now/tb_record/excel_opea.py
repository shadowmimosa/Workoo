import os
import time
from xlwt import Workbook
from xlrd import open_workbook
from openpyxl import load_workbook
from openpyxl import Workbook as init_workbook
from openpyxl.styles import PatternFill


class InvalidOpeaException(Exception):
    """Error for trying to open a non-ooxml file."""


class ExcelOpea(object):
    def __init__(self, use='openpyxl'):
        super().__init__()

        self.use = use

    def set_style(self, name, height, bold=False, color=0):

        style = xlwt.XFStyle()  # 初始化样式
        font = xlwt.Font()  # 为样式创建字体
        font.name = name  # 'Times New Roman'
        font.bold = bold
        font.color_index = color
        font.colour_index = color
        font.height = height

        al = xlwt.Alignment()
        al.horz = 0x01  # 设置左端对齐
        al.vert = 0x01  # 设置垂直居中
        style.alignment = al

        style.font = font

        if color == 0:
            self.style = style
        else:
            self.link_style = style

    def read(self, path, use='openpyxl'):
        if use == 'openpyxl':
            self.read_wkb = load_workbook(path)
            self.read_sheet = self.read_wkb.active

            rows = self.read_sheet.rows
            lines = []

            for row in rows:
                line = [col.value for col in row]
                lines.append(line)

            return lines

        elif use == 'xlrd':
            pass

        else:
            # msg = 'this function only support xlrd and openpyxl to read excel, please check param of use.'
            # raise InvalidOpeaException(msg)
            raise InvalidOpeaException

    def to_excel(self, index, value):
        if self.use == 'openpyxl':
            self.write_sheet.cell(self.row, index + 1, value)
        elif self.use == 'xlwt':
            self.write_sheet.write(self.row, index, value)
        else:
            # msg = 'this function only support xlrd and openpyxl to read excel, please check param of use.'
            # raise InvalidOpeaException(msg)
            raise InvalidOpeaException

    def write(self, content: list or dict):
        if isinstance(content, list) or isinstance(content, tuple):
            for index, value in enumerate(content):
                self.to_excel(index, value)
        elif isinstance(content, dict):
            for index, value in enumerate(self.header):
                try:
                    need_write = content[value]
                except KeyError:
                    need_write = ''
                self.to_excel(index, need_write)
        self.row += 1

    def init_sheet(self, use='openpyxl', header=None):
        self.use = use
        self.header = header

        if use == 'openpyxl':
            self.use = 'openpyxl'
            self.write_wkb = init_workbook()
            self.write_sheet = self.write_wkb.create_sheet('sheet1', 0)
            self.write_sheet.column_dimensions['H'].width = 100.0
            self.row = 1
            self.write(header)

        elif use == 'xlwt':
            self.use = 'xlwt'
            self.write_wkb = Workbook()
            self.write_sheet = self.write_wkb.add_sheet('sheet1',
                                                        cell_overwrite_ok=True)
            self.row = 0
            self.write(header)

        else:
            # msg = 'this function only support xlrd and openpyxl to read excel, please check param of use.'
            # raise InvalidOpeaException(msg)
            raise InvalidOpeaException

    def save(self, path=None, mode='write'):

        if mode == 'write':
            command = 'self.write_wkb.save(path)'
        elif mode == 'read':
            command = 'self.read_wkb.save(path)'

        date = time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())

        if path is None:
            path = '{}.xlsx'.format(
                date) if self.use == 'openpyxl' else '{}.xls'.format(date)
        else:
            if 'xlsx' not in path or 'xls' not in path:
                filename = '{}.xlsx'.format(
                    date) if self.use == 'openpyxl' else '{}.xls'.format(date)
                path = os.path.join(path, filename)
            elif self.use == 'openpyxl' and path[
                    -4:-1] == 'xlsx' or self.use == 'xlwt' and path[
                        -3:-1] == 'xls':
                path = path
            # else:
            #     raise InvalidOpeaException

        eval(command)

        return path

    def append(self, content, fgColor=None):
        self.read_sheet.append(content)

        if fgColor is not None:
            fill = PatternFill(fill_type='solid', fgColor=fgColor)
            cells = self.read_sheet[self.read_sheet._current_row]

            for column in range(len(cells)):
                self.read_sheet.cell(self.read_sheet._current_row,
                                     column + 1).fill = fill


if __name__ == "__main__":
    excel = ExcelOpea()
    excel.read(path='2019-12-16 00-39-15.xlsx')
    excel.append([1, 1, 1, 1])
    excel.save(mode='read')
