from time import time
from os.path import *
from json import dumps
from xlwt import Workbook
from chardet import detect
from openpyxl import load_workbook
from os import chdir, mkdir, listdir
from datetime import datetime, timedelta
from re import compile, search, findall, S


class DealRecord(object):
    def __init__(self):
        self.dirpath = "./data/"
        self.record_path = join(self.dirpath, "聊天记录/")
        self.info_path = join(self.dirpath, "结果/")
        self.ending_path = join(self.dirpath, "话术/")
        self.keyword_path = join(self.dirpath, "关键字/")

        self.header = [
            '客服ID',
            '客户ID',
            '回复时间（秒）',
            '包含行业关键字',
            '有无结尾话术',
            '手机号',
            '手机号所在记录',
            # '聊天记录'
            '访客数据',
        ]
        # self.phone_pattern = compile(r'\d{3}-\d{8}|\d{4}-\{7,8}')
        self.phone_pattern = compile(r'^1[34578]\d{9}')

        self.init_path()
        self.init_sheet()

    def init_path(self):
        try:
            mkdir(self.dirpath)
        except FileExistsError:
            pass

        try:
            mkdir(self.record_path)
        except FileExistsError:
            pass

        try:
            mkdir(self.ending_path)
        except FileExistsError:
            pass

        try:
            mkdir(self.info_path)
        except FileExistsError:
            pass

        try:
            mkdir(self.keyword_path)
        except FileExistsError:
            pass

    def init_sheet(self):
        self.wkb = Workbook()
        self.sheet = self.wkb.add_sheet('sheet1', cell_overwrite_ok=True)
        self.row = 0
        self.write(self.header)

    def save(self):

        self.wkb.save('{}{}.xlsx'.format(self.info_path, int(time() * 1000)))

    def write(self, content: list or dict):
        if isinstance(content, list):
            for index, value in enumerate(content):
                self.sheet.write(self.row, index, value)
        elif isinstance(content, dict):
            for index, value in enumerate(self.header):
                try:
                    need_write = content[value]
                except KeyError:
                    need_write = ''
                self.sheet.write(self.row, index, need_write)

        self.row += 1

    def deal_text(self, content: list):
        time_pattern = compile(
            r'([0-9]{3}[1-9]|[0-9]{2}[1-9][0-9]{1}|[0-9]{1}[1-9][0-9]{2}|[1-9][0-9]{3})-(((0[13578]|1[02])-(0[1-9]|[12][0-9]|3[01]))|((0[469]|11)-(0[1-9]|[12][0-9]|30))|(02-(0[1-9]|[1][0-9]|2[0-8]))) (0[0-9]|1[0-9]|2[0-3]):([0-5][0-9]):([0-5][0-9])'
        )
        record_data = {}
        for item in content:
            if '----------------------------' in item:
                customer = item.replace('----------------------------', '')
                record_data.update({customer: []})
                temp_record = ''
            else:
                sign = search(time_pattern, item)

                if sign:
                    if len(temp_record) == 0:
                        record_data[customer].append(item)
                    else:
                        record_data[customer][-1] = '{}{}'.format(
                            record_data[customer][-1], temp_record)
                        record_data[customer].append(item)
                        temp_record = ''
                else:
                    temp_record = '\n'.join((temp_record, item))

        for record in record_data:
            for index, text in enumerate(record_data[record]):
                temp = text.split('):  ')
                record_data[record][index] = {'text': temp[-1]}
                temp = temp[0].split('(')
                record_data[record][index]['person'] = temp[0]
                record_data[record][index]['time'] = temp[-1].replace(')', '')

        return record_data

    def ending_skill(self, content):
        # for item in reversed(content):

        if content['person'] != self.staff:
            self.info['有无结尾话术'] = '无'
        else:
            if content['text'] not in self.ending_list:
                self.info['有无结尾话术'] = '无'
            else:
                self.info['有无结尾话术'] = '有'

    def deal_keyword(self, content: str):
        exist_list = []
        raw = self.info.get('包含行业关键字')

        if raw is None:
            raw_list = []
        else:
            raw_list = raw.split('\n')

        for keyword in self.keyword_list:
            if keyword in content:
                exist_list.append(keyword)

        exist_list.extend(raw_list)

        if len(exist_list) > 0:
            self.info['包含行业关键字'] = '\n'.join(set(exist_list))

    def deal_trade(self, content: str):

        if self.is_trade is None:
            for trade in self.trade_list.keys():
                if trade in content:
                    self.is_trade = True
                    self.trade_list[trade] += 1

    def tourist(self, date: str):
        self.tourist_second = 60
        date = date.split(' ')[-1].split(':')
        for item in self.tourist_list:
            _time = item[1]

            try:

                seconds = _time.hour * 3600 + _time.minute * 60 + _time.second
                date_seconds = int(date[0]) * 3600 + int(date[1]) * 60 + int(
                    date[2])
            except AttributeError:
                continue
            else:
                if date_seconds - self.tourist_second <= seconds and date_seconds + self.tourist_second >= seconds:
                    self.info['访客数据'] = "{} {} {} {} {}".format(
                        item[0],
                        '{}:{}:{}'.format(item[1].hour, item[1].minute,
                                          item[1].second), item[2], item[3],
                        item[4])

    def get_phone(self, content):
        phone_obj = search(self.phone_pattern, content)

        if phone_obj:
            self.info['手机号'] = phone_obj.group()
            self.info['手机号所在记录'] = content
        elif self.info.get('手机号', None) is None:
            self.info['手机号'] = '无'
            self.info['手机号所在记录'] = '无'

    def deal_time(self, content: list):
        self.is_trade = None
        sign = 0
        record_time = None
        for item in content:
            if item['person'] == self.staff:
                if sign == 0 and record_time is not None:
                    time_diff = str2time(item['time']) - record_time
                    time_temp = self.info.get('回复时间（秒）')

                    if time_temp:
                        if time_diff.total_seconds() > time_temp.total_seconds(
                        ):
                            self.info['回复时间（秒）'] = time_diff
                    else:
                        self.info['回复时间（秒）'] = time_diff

                    print(convert_timedelta(time_diff))
                sign = 1
            elif item['person'] == self.info['客户ID']:
                record_time = str2time(item['time'])

                self.deal_keyword(item['text'])
                self.get_phone(item['text'])
                self.deal_trade(item['text'])

                sign = 0

        if self.info.get('包含行业关键字') is None:
            self.info['包含行业关键字'] = '无'

        if self.info.get('回复时间（秒）') is not None:
            self.info['回复时间（秒）'] = convert_timedelta(self.info['回复时间（秒）'])

            if self.is_trade is not None:
                self.write(self.info)

            elif self.info.get('手机号', None) is not None:
                self.write(self.info)

    def extract_record(self, path):
        with open(path, 'r', encoding='gbk') as fn:
            self.staff = fn.readline().replace('\n', '')
            content = fn.read()

        text = findall(r'={64}.*?={64}(.*?)={64}', content,
                       S)[0].strip('\n').split('\n')
        # text1 = re.search(r'={64}[\n\s\r]*(.*)[\n\s\r]*={64}', content)

        text_list = self.deal_text(text)

        for key, value in text_list.items():
            self.info = {}
            self.info['客服ID'] = self.staff
            self.info['客户ID'] = key
            self.ending_skill(value[-1])
            self.tourist(value[0]['time'])

            self.deal_time(value)

    def main(self):
        self.some_input()

        self.tourist_list = read_xlsx_lines('./data/访客数据/访客数据.xlsx')

        path = '{}结尾话术.txt'.format(self.ending_path)
        self.ending_list = read_lines(path, judge_code(path))

        path = '{}行业关键字.txt'.format(self.keyword_path)
        self.trade_list = read_lines(path, judge_code(path))
        self.trade_list = {i: 0 for i in self.trade_list}

        path = '{}包含关键字列表.txt'.format(self.keyword_path)
        self.keyword_list = read_lines(path, judge_code(path))

        for filename in listdir(self.record_path):
            self.extract_record(join(self.record_path, filename))

        self.save()
        input('文件已保存, 按任意键退出')

    def some_input(self):
        self.tourist_second = check_int(input('请输入访客数据匹配时间, 默认 60 秒\n---> '))

        if not self.tourist_second:
            self.tourist_second = 60


def check_int(item):
    try:
        int(item)
    except TypeError:
        return False
    except ValueError:
        return 60
    else:
        return int(item)


def filter_record(content: str):
    for text in content.split('\n'):
        for line in open('过滤关键字列表.txt', 'r', encoding='utf-8').readlines():
            if '\n' == line:
                continue
            if line.strip('\n').decode('utf-8-sig').encode('gbk') in text:
                content = content.replace(text, '')
    return content


def str2time(string: str):
    return datetime.strptime(string, '%Y-%m-%d %H:%M:%S')


def convert_timedelta(duration: timedelta):
    days, seconds = duration.days, duration.seconds
    hours = days * 24 + seconds // 3600
    minutes = (seconds % 3600) // 60
    seconds = (seconds % 60)

    return '{:0>2d}:{:0>2d}:{:0>2d}'.format(hours, minutes, seconds)


def judge_code(path):
    with open(path, 'rb') as fn:
        data = fn.read()
        charinfo = detect(data)

        return charinfo['encoding']


def read_lines(path, encoding):
    with open(path, 'r', encoding=encoding) as fn:
        lines = fn.readlines()

    return lines


def read_xlsx_lines(path):

    workbook = load_workbook(path)
    booksheet = workbook.active

    rows = booksheet.rows
    lines = []

    for row in rows:
        line = [col.value for col in row]
        # cell_data_1 = booksheet.cell(row=i, column=3).value  #获取第i行1 列的数据
        # cell_data_2 = booksheet.cell(row=i, column=4).value  #获取第i行 2 列的数据
        # cell_data_3 = booksheet.cell(row=i, column=8).value  #获取第i行 3 列的数据
        # cell_data_4 = booksheet.cell(row=i, column=18).value  #获取第i行 4 列的数据
        # print(cell_data_1, cell_data_2, cell_data_3, cell_data_4)
        lines.append(line)
        # print(line)

    return lines


if __name__ == "__main__":
    chdir(dirname(abspath(__file__)))
    DealRecord().main()
