import os
import time
# from xlwt import Workbook
# from xlrd import open_workbook
# import openpyxl
from openpyxl import Workbook
# from openpyxl import load_workbook
# from openpyxl import Workbook as init_workbook
# from openpyxl.styles import PatternFill


class InvalidOpeaException(Exception):
    """Error for trying to open a non-ooxml file."""


class ExcelOpea(object):
    def __init__(self):
        super().__init__()

    def to_excel(self, index, value):
        self.write_sheet.cell(self.row, index + 1, value)

    def write(self, content: list or dict):
        if isinstance(content, list) or isinstance(content, tuple):
            for index, value in enumerate(content):
                self.to_excel(index, value)
        elif isinstance(content, dict):
            for index, value in enumerate(self.header):
                try:
                    need_write = content[value]
                except KeyError:
                    need_write = ''
                self.to_excel(index, need_write)
        self.row += 1

    def init_sheet(self, header=None):
        self.header = header

        self.use = 'openpyxl'
        self.write_wkb = Workbook()
        self.write_sheet = self.write_wkb.create_sheet('sheet1', 0)
        self.row = 1
        self.write(header)

    def save(self, path=None, mode='write'):

        if mode == 'write':
            command = 'self.write_wkb.save(path)'
        elif mode == 'read':
            command = 'self.read_wkb.save(path)'

        date = time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())

        if path is None:
            path = '{}.xlsx'.format(
                date) if self.use == 'openpyxl' else '{}.xls'.format(date)
        else:
            if 'xlsx' not in path or 'xls' not in path:
                filename = '{}.xlsx'.format(
                    date) if self.use == 'openpyxl' else '{}.xls'.format(date)
                path = os.path.join(path, filename)
            elif self.use == 'openpyxl' and path[
                    -4:-1] == 'xlsx' or self.use == 'xlwt' and path[
                        -3:-1] == 'xls':
                path = path
            # else:
            #     raise InvalidOpeaException

        eval(command)

        return path


if __name__ == "__main__":
    excel = ExcelOpea()
    excel.read(path='2019-12-16 00-39-15.xlsx')
    excel.append([1, 1, 1, 1])
    excel.save(mode='read')
