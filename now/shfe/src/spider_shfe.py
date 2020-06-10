import os
import re
import json
import arrow
import requests
from openpyxl import Workbook as init_workbook


class ExcelOpea(object):
    def __init__(self, header):
        self.header = header
        self.wkb = init_workbook()
        super().__init__()

    def write(self, content, sheet_name, init=False):
        # index = self.wkb.sheetnames.index(sheet_name)
        sheet = self.wkb[sheet_name]
        row = sheet.max_row + 1 if init is False else 1
        for index, value in enumerate(content):
            sheet.cell(row, index + 1, value)

    def create_sheet(self, sheet_name='sheet1'):
        self.wkb.create_sheet(sheet_name, 0)
        self.write(self.header, sheet_name, init=True)

    def save(self, name):
        self.wkb.save(f'./data/{name}.xlsx')


def remove(content):
    return content.replace(' ', '')


def is_leap(year: int):
    if ((year % 4 == 0 and year % 100 != 0) or (year % 400 == 0)):
        days = 366
    else:
        days = 365

    return days


def get_date(years: list = [arrow.now().year]):
    for year in years:
        count = 0
        days_count = is_leap(year)
        start_date = f'{year}-1-1'
        while count < days_count:
            date = arrow.get(start_date).shift(days=count).format('YYYYMMDD')
            count += 1
            yield date


class DealShfe(object):
    def __init__(self, min_year, max_year):
        self.min_year = min_year
        self.max_year = max_year
        self.sheet_name = ['铜期权', '黄金期权', '天胶期权']
        self.header = {
            'Accept': 'application/json, text/javascript, */*; q=0.01',
            'User-Agent':
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/79.0.3945.130 Safari/537.36',
            'X-Requested-With': 'XMLHttpRequest',
            'Referer':
            'http://www.shfe.com.cn/statements/dataview.html?paramid=pm',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9',
        }
        self.create_sheet()

    def create_sheet(self):
        column = [
            '合约代码', '开盘价', '最高价', '最低价', '收盘价', '前结算价', '结算价', '涨跌1', '涨跌2',
            '成交量', '持仓量', '持仓量变化', '成交额', '德尔塔Delta', '行权量'
        ]

        self.excel = ExcelOpea(column)
        for sheet in self.sheet_name:
            self.excel.create_sheet(sheet)

    def run(self, date):
        uri = f'http://www.shfe.com.cn/data/dailydata/option/kx/kx{date}.dat'
        resp = requests.get(uri, headers=self.header)

        if resp.status_code != 200:
            return

        data = json.loads(remove(resp.text))

        for line in data.get('o_curinstrument', []):
            info = []
            info.append(line.get('INSTRUMENTID'))
            info.append(line.get('OPENPRICE'))
            info.append(line.get('HIGHESTPRICE'))
            info.append(line.get('LOWESTPRICE'))
            info.append(line.get('CLOSEPRICE'))
            info.append(line.get('PRESETTLEMENTPRICE'))
            info.append(line.get('SETTLEMENTPRICE'))
            info.append(line.get('ZD1_CHG'))
            info.append(line.get('ZD2_CHG'))
            info.append(line.get('VOLUME'))
            info.append(line.get('OPENINTEREST'))
            info.append(line.get('OPENINTERESTCHG'))
            info.append(round(line.get('TURNOVER'),
                              2)) if line.get('TURNOVER') else info.append(
                                  line.get('TURNOVER'))
            info.append(round(line.get('DELTA'),
                              2)) if line.get('DELTA') else info.append(
                                  line.get('DELTA'))
            info.append(line.get('EXECVOLUME'))

            product = line.get('PRODUCTNAME')
            self.excel.write(info,
                             product) if product in self.sheet_name else True
        self.excel.save(date)
        print(f'{date} 已保存')

    def main(self):
        date = get_date([x for x in range(self.min_year, self.max_year + 1)])

        while True:
            try:
                now = next(date)
            except StopIteration:
                break
            else:
                try:
                    self.run(now)
                except Exception as exc:
                    print(f'{now} 保存失败 - {exc}')


def main():
    min_year = 2019
    max_year = 2020

    spider = DealShfe(min_year, max_year)
    try:
        spider.main()
    except Exception as exc:
        print(f'Something Wrong, error is {exc}')


if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    main()