import os
import re
import time
from openpyxl import Workbook

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')


class InvalidOpeaException(Exception):
    """Error for trying to open a non-ooxml file."""


class ExcelOpea(object):
    def __init__(self):
        self.use = 'openpyxl'
        self.write_wkb = Workbook()
        super().__init__()

    def to_excel(self, index, value):
        value = ILLEGAL_CHARACTERS_RE.sub(r'', value)

        self.write_sheet.cell(self.row, index + 1, value)

    def write(self, content: list or dict, sheet_name=None):
        if sheet_name is not None:
            self.write_sheet = self.write_wkb[sheet_name]

        if isinstance(content, list) or isinstance(content, tuple):
            for index, value in enumerate(content):
                self.to_excel(index, value)
        elif isinstance(content, dict):
            for index, value in enumerate(self.header):
                try:
                    need_write = content[value]
                except KeyError:
                    need_write = ''
                self.to_excel(index, need_write)
        self.row += 1

    def write_many(self, content: list, sheet_name=None):
        for item in content:
            self.write(item, sheet_name)

    def init_sheet(self, header=None, sheet_name='sheet1'):
        self.header = header
        self.write_sheet = self.write_wkb.create_sheet(sheet_name, 0)
        self.row = 1
        self.write(header)

    def save(self, path=None, mode='write'):

        if mode == 'write':
            command = 'self.write_wkb.save(path)'
        elif mode == 'read':
            command = 'self.read_wkb.save(path)'

        date = time.strftime("%Y-%m-%d %H-%M-%S", time.localtime())

        if path is None:
            path = '{}.xlsx'.format(
                date) if self.use == 'openpyxl' else '{}.xls'.format(date)
        else:
            if 'xlsx' not in path or 'xls' not in path:
                filename = '{}.xlsx'.format(
                    date) if self.use == 'openpyxl' else '{}.xls'.format(date)
                path = os.path.join(path, filename)
            elif self.use == 'openpyxl' and path[
                    -4:-1] == 'xlsx' or self.use == 'xlwt' and path[
                        -3:-1] == 'xls':
                path = path
        path = path.replace('*', '')
        eval(command)

        return path


if __name__ == "__main__":
    excel = ExcelOpea()
    excel.read(path='2019-12-16 00-39-15.xlsx')
    excel.append([1, 1, 1, 1])
    excel.save(mode='read')
